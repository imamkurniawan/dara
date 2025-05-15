# pip install pandas, openpyxl, plotly

from flask import Flask, redirect, request, render_template, url_for, flash, session, jsonify, send_file
import pandas as pd
import numpy as np
import plotly.graph_objs as go
import plotly.io as pio
import arrow # untuk formating date
from openpyxl import load_workbook
import os
import csv
import logging
from datetime import datetime, date
import requests
# import subprocess
import re
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.naive_bayes import MultinomialNB
from sklearn.pipeline import Pipeline
import math
from io import BytesIO

app = Flask(__name__)
app.secret_key = 'supersecretkey'  # Untuk keperluan flash message

############################## Config Host ###################################
##############################################################################
host = 'http://localhost:5000/'

# Penyimpanan sementara untuk data user (gunakan database untuk produksi)
users = [
    {
    'username':'admin@dara.com',
    'password':'admin',
    'status' : 'Administrator'
    },
    {
    'username':'operator@dara.com',
    'password':'operator',
    'status' : 'Operator'
    },
    {
    'username':'user@dara.com',
    'password':'user',
    'status' : 'User'
    }
]

############################ Config Dataset ##################################
##############################################################################

dataset = {
    # File input
    'file_ulasan':'dataset/gmaps_review.xlsx',
    'file_pengaduan':'dataset/pengaduan.xlsx',
    'data_unit':'dataset/data_unit.xlsx',

    # File proses
    'daftar_person':'dataset/persons.txt',
    'daftar_place':'dataset/places.txt',
    'data_training':"dataset/training_data.xlsx",

    # File Hasil
    'results_persons_ulasan':'dataset/persons_ulasan_results.csv',
    'results_places_ulasan':'dataset/places_ulasan_results.csv',
    'results_persons_pengaduan':'dataset/persons_pengaduan_results.csv',
    'results_places_pengaduan':'dataset/places_pengaduan_results.csv',
    
    'results_cleaned_ulasan':'dataset/results_cleaned_ulasan.csv',
    'results_prediction_ulasan':'dataset/results_prediction_ulasan.csv',
    'results_group_prediction_ulasan':'dataset/results_group_prediction_ulasan.csv',
    'results_transformed_prediction_ulasan':'dataset/results_transformed_prediction_ulasan.csv',

    'results_cleaned_pengaduan':'dataset/results_cleaned_pengaduan.csv',
    'results_prediction_pengaduan':'dataset/results_prediction_pengaduan.csv',
    'results_group_prediction_pengaduan':'dataset/results_group_prediction_pengaduan.csv',
    'results_transformed_prediction_pengaduan':'dataset/results_transformed_prediction_pengaduan.csv',

    # file faq
    'file_faq':'dataset/faq.xlsx',

    # file log pengunjung
    'log_file':'dataset/pengunjung.csv'

}

#### Fungsi rekam log pengunjung ####
# Matikan log bawaan Flask (opsional)
log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)

log_file = dataset['log_file']

#if not os.path.exists(log_file):
#    open(log_file, 'w').close()

# Buat file dan header kalau belum ada
if not os.path.exists(log_file):
    with open(log_file, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(['waktu', 'ip', 'url', 'referer', 'user_agent'])

def log_info_pengunjung():
    path = request.path

    # Abaikan request file statis
    if path.startswith('/static') or path.startswith('/js') or path.endswith(('.css', '.js', '.jpg', '.png', '.woff2', '.ico')):
        return

    waktu = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ip = request.headers.get('X-Forwarded-For', request.remote_addr)
    url = request.url
    referer = request.referrer or 'Langsung (direct)'
    user_agent = request.headers.get('User-Agent', 'Unknown')

    with open(log_file, mode='a', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow([waktu, ip, url, referer, user_agent])


#### fungsi ubah date menjadi durasi (cth: '3 hari yang lalu, seminggu yg lalu,') ####
def humanize_date(date):
    return arrow.get(date).humanize(locale='id')  # Bahasa Indonesia
# Tambahkan filter ke Jinja
app.jinja_env.filters['humanize_date'] = humanize_date

#### Fungsi higlight word (untup person dan pace) ####
@app.template_filter('highlight_replace')
def highlight_replace(text, word):
    text = text.lower()
    return text.replace(word, f"<span class='yellow_color'>{word}</span>")
# Registrasi filter
# app.jinja_env.filters['highlight_replace'] = highlight_replace

#####################################
# Fungsi untuk membaca data pengaduan
def get_data_pengaduan():
    df = pd.read_excel(dataset['file_pengaduan'])
    # Ekstrak year
    # df['thn'] = df['tgl_pengaduan'].dt.year
    df['thn_pengaduan'] = df['tgl_pengaduan'].dt.year
    # Ekstrak bulsn
    df['bln'] = df['tgl_pengaduan'].dt.month
    df.sort_values(by="reviewer_id", ascending=False, inplace=True)
    return df

def get_last_record_pengaduan():
    df = get_data_pengaduan()
    last_record = df['record_no'].max()
    return last_record

#####################################
def calculate_sentiment():
    # ambil file transformed prediction
    df = pd.read_csv(dataset['results_transformed_prediction_ulasan'])
    total_sentimen = len(df)
    df_positif = df.query("predicted_sentiment == 'positif'")
    total_positif = len(df_positif)
    persen_positif = int(round((total_positif/total_sentimen)*100,0))
    df_negatif = df.query("predicted_sentiment == 'negatif'")
    total_negatif = len(df_negatif)
    persen_negatif = int(round((total_negatif/total_sentimen)*100,0))
    df_biasa = df.query("predicted_sentiment == 'biasa'")
    total_biasa = len(df_biasa)
    persen_biasa = int(round((total_biasa/total_sentimen)*100,0))
    data = {
        'total_sentimen':total_sentimen,
        'total_positif':total_positif,
        'total_negatif':total_negatif,
        'total_biasa':total_biasa,
        'persen_positif':persen_positif,
        'persen_negatif':persen_negatif,
        'persen_biasa':persen_biasa
    }
    return(data)

##################################
# Fungsi untuk membaca data ulasan
def get_data_ulasan():
    # Baca file Excel
    df = pd.read_excel(dataset['file_ulasan'])
    # Pilih kolom yang ingin ditampilkan
    selected_columns = ["record_no","reviewer_id","name","link","thumbnail","reviews","photos","localGuide","rating","duration", "tgl_ulasan", "snippet", "fixed_review", "likes", "thn_ulasan"]  # Ganti dengan nama kolom yang ingin ditampilkan
    df = df[selected_columns]
    # Konversi dengan format yang benar
    df['tgl_ulasan'] = pd.to_datetime(df['tgl_ulasan'], dayfirst=True)
    # Ekstrak year
    df['thn'] = df['tgl_ulasan'].dt.year
    df.sort_values(by="record_no", ascending=False, inplace=True)
    return df

def get_last_record():
    df = get_data_ulasan()
    last_record = df['record_no'].max()
    return last_record

def calculate_review_stats(df):
    total_reviews = len(df)
    average_rating = df['rating'].mean()
    average_rating = round(average_rating,2)
    return total_reviews, average_rating

def get_ratings_per_year_and_value(df):
    ratings_per_year_value = df.groupby(['thn', 'rating']).size().reset_index(name='count')
    pivot_table = ratings_per_year_value.pivot(index='thn', columns='rating', values='count').fillna(0)
    return pivot_table

def get_total_ulasan_tahunan():
    df = get_data_ulasan()
    # Group by berdasarkan tahun dan hitung jumlah record
    result = df.groupby('thn').size().reset_index(name='jumlah_records')
    return result

def get_total_rating():
    df = get_data_ulasan()
    # Group by berdasarkan tahun dan hitung jumlah record
    result = df.groupby('rating').size().reset_index(name='jumlah_records')
    return result

def load_and_group_data(thn='semua',topik='semua',sentimen='semua'):
    # Baca tabel utama dan tambahan
    df1 = pd.read_excel(dataset['file_ulasan'])
    if thn != 'semua':
        df1 = df1.query("thn_ulasan == @thn")
    # Baca tabel transformed
    df2 = pd.read_csv(dataset['results_transformed_prediction_ulasan'])
    if thn != 'semua':
        df2 = df2.query("thn == @thn")
    if topik != 'semua':
        df2 = df2.query("predicted_topic == @topik")
    if sentimen != 'semua':
        df2 = df2.query("predicted_sentiment == @sentimen")
    
    persons_df = pd.read_csv(dataset['results_persons_ulasan'])
    places_df = pd.read_csv(dataset['results_places_ulasan'])    
    # Gabungkan data
    merged_data = df1.merge(df2, on=['record_no', 'reviewer_id'], how='right')
    merged_data = merged_data.merge(persons_df, on='reviewer_id', how='left')
    merged_data = merged_data.merge(places_df, on='reviewer_id', how='left')
    print(merged_data)
    # Filter snippet yang kosong
    merged_data = merged_data[merged_data['snippet'].notna() & (merged_data['snippet'] != '')]
    # Hapus duplikasi data berdasarkan semua kolom
    merged_data = merged_data.drop_duplicates()    
    # Kelompokkan berdasarkan record_no dan reviewer_id
    grouped_data = []
    # Urutkan data berdasarkan record_no secara descending
    for (record_no, reviewer_id), group in merged_data.groupby(['record_no', 'reviewer_id']):
        main_data = {
            'record_no': record_no,
            'reviewer_id': reviewer_id,
            'name': group.iloc[0]['name'],
            'link': group.iloc[0]['link'],
            'localGuide': group.iloc[0]['localGuide'],
            'rating': group.iloc[0]['rating'],
            'snippet': group.iloc[0]['snippet'],
            'fixed_review': group.iloc[0]['fixed_review'],
            'likes': group.iloc[0]['likes'],
            'tgl_ulasan': group.iloc[0]['tgl_ulasan'],
            'thn_ulasan': group.iloc[0]['thn_ulasan']
        }
        # Data prediksi (hilangkan duplikasi)
        predicted_data = group[['predicted_topic', 'predicted_sentiment', 'count']].drop_duplicates().to_dict(orient='records')
        # Data persons (hilangkan duplikasi)
        persons_data = group[['persons', 'persons_count']].drop_duplicates().dropna().to_dict(orient='records')
        # Data places (hilangkan duplikasi)
        places_data = group[['places', 'places_count']].drop_duplicates().dropna().to_dict(orient='records')
        
        grouped_data.append({
            'main_data': main_data,
            'predicted_data': predicted_data,
            'persons_data': persons_data,
            'places_data': places_data
        })
    # (Opsional) Urutkan grouped_data berdasarkan record_no
    grouped_data = sorted(grouped_data, key=lambda x: x['main_data']['record_no'], reverse=True)
    # Debug: Cek urutan setelah grouped_data dibuat
    # print("Grouped data record_no (desc):", [item['main_data']['record_no'] for item in grouped_data])    
    return grouped_data

def load_and_group_data_pengaduan(thn='semua',topik='semua',status='semua'):
    # Baca tabel utama dan tambahan
    # df1 = pd.read_excel(dataset['file_pengaduan'])
    df1 = get_data_pengaduan()
    print (df1)
    if thn != 'semua':
        df1 = df1.query("thn_pengaduan == @thn")
    if status != 'semua':
        df1 = df1.query("status_pengaduan == @status")
    # Baca tabel transformed
    df2 = pd.read_csv(dataset['results_transformed_prediction_pengaduan'])
    if thn != 'semua':
        df2 = df2.query("thn == @thn")
    if topik != 'semua':
        df2 = df2.query("predicted_topic == @topik")
    
    persons_df = pd.read_csv(dataset['results_persons_pengaduan'])
    #print (persons_df)
    places_df = pd.read_csv(dataset['results_places_pengaduan'])    
    # Gabungkan data
    merged_data = df1.merge(df2, on=['record_no', 'reviewer_id'], how='right')
    merged_data = merged_data.merge(persons_df, on='reviewer_id', how='left')
    merged_data = merged_data.merge(places_df, on='reviewer_id', how='left')
    
    # Filter snippet yang kosong
    merged_data = merged_data[merged_data['fixed_pengaduan'].notna() & (merged_data['fixed_pengaduan'] != '')]
    # Hapus duplikasi data berdasarkan semua kolom
    merged_data = merged_data.drop_duplicates()
    print(merged_data)    
    # Kelompokkan berdasarkan record_no dan reviewer_id
    grouped_data = []
    # Urutkan data berdasarkan record_no secara descending
    for (record_no, reviewer_id), group in merged_data.groupby(['record_no', 'reviewer_id']):
        main_data = {
            'record_no': record_no,
            'reviewer_id': reviewer_id,
            'nama': group.iloc[0]['nama'],
            'sumber': group.iloc[0]['sumber'],
            'fixed_pengaduan': group.iloc[0]['fixed_pengaduan'],
            'solusi_pengaduan': group.iloc[0]['solusi_pengaduan'],
            'status_pengaduan': group.iloc[0]['status_pengaduan'],
            'tgl_pengaduan': group.iloc[0]['tgl_pengaduan'],
            'thn_pengaduan': group.iloc[0]['thn_pengaduan']
        }
        # Data prediksi (hilangkan duplikasi)
        predicted_data = group[['predicted_topic', 'predicted_sentiment', 'count']].drop_duplicates().to_dict(orient='records')
        # Data persons (hilangkan duplikasi)
        persons_data = group[['persons', 'persons_count']].drop_duplicates().dropna().to_dict(orient='records')
        # Data places (hilangkan duplikasi)
        places_data = group[['places', 'places_count']].drop_duplicates().dropna().to_dict(orient='records')
        
        grouped_data.append({
            'main_data': main_data,
            'predicted_data': predicted_data,
            'persons_data': persons_data,
            'places_data': places_data,
        })
    # (Opsional) Urutkan grouped_data berdasarkan record_no
    grouped_data = sorted(grouped_data, key=lambda x: x['main_data']['record_no'], reverse=True)
    # Debug: Cek urutan setelah grouped_data dibuat
    # print("Grouped data record_no (desc):", [item['main_data']['record_no'] for item in grouped_data])    
    return grouped_data

####################################################### HTML Templates Backend
##############################################################################
# Halaman untuk menangani error
@app.errorhandler(Exception)
def handle_exception(e):
    return render_template("error.html", error=e), 500

# jalankan dulu sebelum request
@app.before_request
def sebelum_request():
    # Tangani hasil return dari cek_login()
    redirect_response = cek_login()
    if redirect_response:
        return redirect_response  # Jika ada redirect, kembalikan response tersebut
    
    # tulis log
    log_info_pengunjung()

# cek sudah login apa belum
def cek_login():
    allowed_routes = ['login', 'static', 'lacak_pengaduan']  # Route yang diizinkan tanpa login
    if request.endpoint not in allowed_routes and not session.get('user'):
        return redirect(url_for('login'))

@app.route('/login',  methods=['GET','POST'])
def login():
    pesan = ''
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        user = next((user for user in users if user['username'] == username and user['password'] == password), None)
        if user:
            print(f"Login berhasil! Selamat datang, {user['status']}.")
            session['user'] = username
            session['status'] = user['status']
            return redirect(url_for('index'))
        else:
            pesan = "Login gagal! Username atau password salah."

    return render_template("login.html",pesan=pesan)

@app.route('/logout')
def logout():
    session.pop('user', None)
    session.pop('status', None)
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
def index():

    # Muat dan kelompokkan data ulasan
    df = get_data_ulasan()
    data = df.iloc[:4]
    # Konversi DataFrame ke daftar untuk dikirim ke template
    data = data.to_dict(orient="records")  # Mengonversi setiap baris menjadi dictionary

    # Muat data pengaduan
    df_pengaduan = get_data_pengaduan()
    jml_pengaduan = len(df_pengaduan)
    df_pengaduan_proses = df_pengaduan.query("status_pengaduan == 'selesai'")
    jml_pengaduan_proses = len(df_pengaduan_proses)
    df_pengaduan_selesai = df_pengaduan.query("status_pengaduan == 'selesai'")
    jml_pengaduan_selesai = len(df_pengaduan_selesai)
    stat_pengaduan = {
        "jml_pengaduan":jml_pengaduan,
        "jml_pengaduan_proses":jml_pengaduan_proses,
        "jml_pengaduan_selesai":jml_pengaduan_selesai
    }

    data_pengaduan = df_pengaduan.iloc[:4]
    data_pengaduan = data_pengaduan.to_dict(orient="records")

    # ambil statistik ulasan
    stat = calculate_review_stats(df)
    sentimen = calculate_sentiment()    
    return render_template('index.html', data=data, stat=stat, sentimen=sentimen, data_pengaduan=data_pengaduan, stat_pengaduan=stat_pengaduan)

@app.route('/overview')
def overview():
    # ambil total review dan rata-rata rating
    data_ulasan = get_data_ulasan()
    total_reviews, average_rating = calculate_review_stats(data_ulasan)
    
    data_pengaduan = get_data_pengaduan()
    total_pengaduan_internal = len(data_pengaduan)
    print (data_pengaduan)
    
    # ambil tahun sekarang
    current_date = date.today()
    thn = current_date.year
    data_pengaduan_thn_terkini = data_pengaduan.query("thn_pengaduan == @thn")
    total_pengaduan_internal_terkini = len(data_pengaduan_thn_terkini)
    print (data_pengaduan_thn_terkini)

    data_pengaduan_selesai = data_pengaduan.query("status_pengaduan == 'selesai'")
    total_pengaduan_selesai = len(data_pengaduan_selesai)

    data_pengaduan_on_proses = data_pengaduan.query("status_pengaduan == 'proses'")
    total_pengaduan_on_proses = len(data_pengaduan_on_proses)

    # ambil sentimen
    sentimen = calculate_sentiment()

    data = {
        'total_reviews':total_reviews,
        'average_rating':average_rating,
        'total_pengaduan_internal':total_pengaduan_internal,
        'total_pengaduan_internal_terkini':total_pengaduan_internal_terkini,
        'total_pengaduan_selesai':total_pengaduan_selesai,
        'total_pengaduan_on_proses':total_pengaduan_on_proses,
        'sentimen':sentimen
    }
    return render_template('mydash.html', data=data, thn=thn)

@app.route('/pengaduan')
def pengaduan():
    current_date = date.today()
    thn = current_date.year
    return render_template('filter_pengaduan.html',thn=thn)

@app.route('/pengaduan_lengkap', methods=["GET","POST"])
def pengaduan_lengkap():
    if request.method != "POST":
        return redirect('/pengaduan')
    else:
        if request.form.get('tahun') != 'semua':
            thn_ulasan = int(request.form.get('tahun'))
        else:
            thn_ulasan = 'semua'
        topik = request.form.get('topik')
        status = request.form.get('status')
        
        # Muat dan kelompokkan data
        data = load_and_group_data_pengaduan(thn_ulasan, topik, status)
        count = len(data)
        filter = {
            'thn': thn_ulasan,
            'topik': topik,
            'status': status,
            'count':count
        }
    #    print (data)
        return render_template('pengaduan_with_topic.html', data=data, filter=filter)
    #    return f'tes {thn_ulasan}{topik}{status}{data}'

@app.route('/persons_pengaduan')
def persons_pengaduan():
    df = pd.read_csv(dataset['results_persons_pengaduan'])
    df = df.groupby(["persons"]).size().reset_index(name='count')
    df.sort_values(by="count", ascending=False, inplace=True)
    data = df.to_dict(orient="records")
    # print(data)
    return render_template("persons_pengaduan.html",data=data)

@app.route("/person_pengaduan_review",methods=["GET","POST"])
def person_pengaduan_review():
    person = request.args.get('person')
    df = pd.read_csv(dataset['results_persons_pengaduan'])
    df = df.query("persons == @person")
    # Ambil reviewer_id dari filtered_df
    reviewer_ids = df['reviewer_id'].unique()
    
    # Baca dataset pengaduan
    df_reviews = pd.read_excel(dataset['file_pengaduan'])
    # df_reviews.sort_values(by="record_no", ascending=False, inplace=True)    
    
    # Filter df_reviews berdasarkan reviewer_id
    filtered_reviews = df_reviews[df_reviews['reviewer_id'].isin(reviewer_ids)]
    
    # ambil topic
    # topic_reviews = pd.read_csv('dataset/transformed_review_prediction.csv')
    topic_reviews = pd.read_csv(dataset['results_transformed_prediction_pengaduan'])
    filtered_topic = topic_reviews[topic_reviews['reviewer_id'].isin(reviewer_ids)]

    merged_data = filtered_reviews.merge(filtered_topic, on=['record_no', 'reviewer_id'], how='left')    
    # merged_data.sort_values(by="record_no", ascending=False, inplace=True)    
    # Kelompokkan berdasarkan record_no dan reviewer_id
    grouped_data = []
    # Urutkan data berdasarkan record_no secara descending
    for (record_no, reviewer_id), group in merged_data.groupby(['record_no', 'reviewer_id']):
        main_data = {
            'record_no': record_no,
            'reviewer_id': reviewer_id,
            'nama': group.iloc[0]['nama'],
            'sumber': group.iloc[0]['sumber'],
            'fixed_pengaduan': group.iloc[0]['fixed_pengaduan'],
            'status_pengaduan': group.iloc[0]['status_pengaduan'],
            'tgl_pengaduan': group.iloc[0]['tgl_pengaduan'],
            'thn_pengaduan': group.iloc[0]['thn']
     }        
        predicted_data = group[['record_no','predicted_topic', 'predicted_sentiment', 'count']].drop_duplicates().to_dict(orient='records')
        grouped_data.append({
                'main_data': main_data,
                'topic': predicted_data
            })    
    # Sort berdasarkan record_no descending
    grouped_data = sorted(grouped_data, key=lambda x: x['main_data']['record_no'], reverse=True)
    return render_template("person_pengaduan_review.html",data=grouped_data, person=person)

@app.route('/places_pengaduan')
def places_pengaduan():
    df = pd.read_csv(dataset['results_places_pengaduan'])
    df = df.groupby(["places"]).size().reset_index(name='count')
    df.sort_values(by="count", ascending=False, inplace=True)
    data = df.to_dict(orient="records")
    # print(data)
    return render_template("places_pengaduan.html",data=data)

@app.route("/place_pengaduan_review",methods=["GET","POST"])
def place_pengaduan_review():
    place = request.args.get('place')
    df = pd.read_csv(dataset['results_places_pengaduan'])
    df = df.query("places == @place")
    # Ambil reviewer_id dari filtered_df
    reviewer_ids = df['reviewer_id'].unique()
    
    # Baca dataset pengaduan
    df_reviews = pd.read_excel(dataset['file_pengaduan'])
    # df_reviews.sort_values(by="record_no", ascending=False, inplace=True)    
    
    # Filter df_reviews berdasarkan reviewer_id
    filtered_reviews = df_reviews[df_reviews['reviewer_id'].isin(reviewer_ids)]
    
    # ambil topic
    # topic_reviews = pd.read_csv('dataset/transformed_review_prediction.csv')
    topic_reviews = pd.read_csv(dataset['results_transformed_prediction_pengaduan'])
    filtered_topic = topic_reviews[topic_reviews['reviewer_id'].isin(reviewer_ids)]

    merged_data = filtered_reviews.merge(filtered_topic, on=['record_no', 'reviewer_id'], how='left')    
    # merged_data.sort_values(by="record_no", ascending=False, inplace=True)    
    # Kelompokkan berdasarkan record_no dan reviewer_id
    grouped_data = []
    # Urutkan data berdasarkan record_no secara descending
    for (record_no, reviewer_id), group in merged_data.groupby(['record_no', 'reviewer_id']):
        main_data = {
            'record_no': record_no,
            'reviewer_id': reviewer_id,
            'nama': group.iloc[0]['nama'],
            'sumber': group.iloc[0]['sumber'],
            'fixed_pengaduan': group.iloc[0]['fixed_pengaduan'],
            'status_pengaduan': group.iloc[0]['status_pengaduan'],
            'tgl_pengaduan': group.iloc[0]['tgl_pengaduan'],
            'thn_pengaduan': group.iloc[0]['thn']
     }        
        predicted_data = group[['record_no','predicted_topic', 'predicted_sentiment', 'count']].drop_duplicates().to_dict(orient='records')
        grouped_data.append({
                'main_data': main_data,
                'topic': predicted_data
            })    
    # Sort berdasarkan record_no descending
    grouped_data = sorted(grouped_data, key=lambda x: x['main_data']['record_no'], reverse=True)
    return render_template("place_pengaduan_review.html",data=grouped_data, place=place)


@app.route('/filter_ulasan', methods=["GET","POST"])
def filter_ulasan():
    return render_template('filter_reviews.html')

@app.route('/ulasan_lengkap', methods=["GET","POST"])
def ulasan_lengkap():
    if request.method != "POST":
        return redirect('/filter_ulasan')
    else:
        if request.form.get('tahun') != 'semua':
            thn_ulasan = int(request.form.get('tahun'))
        else:
            thn_ulasan = 'semua'
        topik = request.form.get('topik')
        sentimen = request.form.get('sentimen')
        
        # Muat dan kelompokkan data
        data = load_and_group_data(thn_ulasan, topik, sentimen)
        count = len(data)
        filter = {
            'thn': thn_ulasan,
            'topik': topik,
            'sentimen': sentimen,
            'count':count
        }
        return render_template('reviews_with_topic.html', data=data, filter=filter)

@app.route('/persons')
def persons():
    df = pd.read_csv(dataset['results_persons_ulasan'])
    df = df.groupby(["persons"]).size().reset_index(name='count')
    df.sort_values(by="count", ascending=False, inplace=True)
    data = df.to_dict(orient="records")
    # print(data)
    return render_template("persons.html",data=data)

@app.route("/person_review",methods=["GET","POST"])
def person_review():
    person = request.args.get('person')
    df = pd.read_csv(dataset['results_persons_ulasan'])
    df = df.query("persons == @person")
    # Ambil reviewer_id dari filtered_df
    reviewer_ids = df['reviewer_id'].unique()
    
    # Baca dataset 'gmaps_review.xlsx'
    df_reviews = pd.read_excel('dataset/gmaps_review.xlsx')
    # df_reviews.sort_values(by="record_no", ascending=False, inplace=True)    
    
    # Filter df_reviews berdasarkan reviewer_id
    filtered_reviews = df_reviews[df_reviews['reviewer_id'].isin(reviewer_ids)]
    
    # ambil topic
    # topic_reviews = pd.read_csv('dataset/transformed_review_prediction.csv')
    topic_reviews = pd.read_csv(dataset['results_transformed_prediction_ulasan'])
    filtered_topic = topic_reviews[topic_reviews['reviewer_id'].isin(reviewer_ids)]

    merged_data = filtered_reviews.merge(filtered_topic, on=['record_no', 'reviewer_id'], how='left')    
    # merged_data.sort_values(by="record_no", ascending=False, inplace=True)    
    # Kelompokkan berdasarkan record_no dan reviewer_id
    grouped_data = []
    # Urutkan data berdasarkan record_no secara descending
    for (record_no, reviewer_id), group in merged_data.groupby(['record_no', 'reviewer_id']):
        main_data = {
            'record_no': record_no,
            'reviewer_id': reviewer_id,
            'name': group.iloc[0]['name'],
            'link': group.iloc[0]['link'],
            'localGuide': group.iloc[0]['localGuide'],
            'rating': group.iloc[0]['rating'],
            'snippet': group.iloc[0]['snippet'],
            'fixed_review': group.iloc[0]['fixed_review'],
            'likes': group.iloc[0]['likes'],
            'tgl_ulasan': group.iloc[0]['tgl_ulasan'],
            'thn_ulasan': group.iloc[0]['thn_ulasan']
     }        
        predicted_data = group[['record_no','predicted_topic', 'predicted_sentiment', 'count']].drop_duplicates().to_dict(orient='records')
        grouped_data.append({
                'main_data': main_data,
                'topic': predicted_data
            })    
    # Sort berdasarkan record_no descending
    grouped_data = sorted(grouped_data, key=lambda x: x['main_data']['record_no'], reverse=True)
    return render_template("person_review.html",data=grouped_data, person=person)

@app.route('/places')
def places():
    df = pd.read_csv(dataset['results_places_ulasan'])
    df = df.groupby(["places"]).size().reset_index(name='count')
    df.sort_values(by="count", ascending=False, inplace=True)
    data = df.to_dict(orient="records")
    # print(data)
    return render_template("places.html",data=data)

@app.route("/place_review",methods=["GET","POST"])
def place_review():
    place = request.args.get('place')
    df = pd.read_csv(dataset['results_places_ulasan'])
    df = df.query("places == @place")
    # Ambil reviewer_id dari filtered_df
    reviewer_ids = df['reviewer_id'].unique()

    # Baca dataset 'gmaps_review.xlsx'
    df_reviews = pd.read_excel(dataset['file_ulasan'])
    # df_reviews.sort_values(by="record_no", ascending=False, inplace=True)    
    
    # Filter df_reviews berdasarkan reviewer_id
    filtered_reviews = df_reviews[df_reviews['reviewer_id'].isin(reviewer_ids)]
    
    # ambil topic
    # topic_reviews = pd.read_csv('dataset/transformed_review_prediction.csv')
    topic_reviews = pd.read_csv(dataset['results_transformed_prediction_ulasan'])
    filtered_topic = topic_reviews[topic_reviews['reviewer_id'].isin(reviewer_ids)]

    merged_data = filtered_reviews.merge(filtered_topic, on=['record_no', 'reviewer_id'], how='left')    
    # merged_data.sort_values(by="record_no", ascending=False, inplace=True)    
    # Kelompokkan berdasarkan record_no dan reviewer_id
    grouped_data = []
    # Urutkan data berdasarkan record_no secara descending
    for (record_no, reviewer_id), group in merged_data.groupby(['record_no', 'reviewer_id']):
        main_data = {
            'record_no': record_no,
            'reviewer_id': reviewer_id,
            'name': group.iloc[0]['name'],
            'link': group.iloc[0]['link'],
            'localGuide': group.iloc[0]['localGuide'],
            'rating': group.iloc[0]['rating'],
            'snippet': group.iloc[0]['snippet'],
            'fixed_review': group.iloc[0]['fixed_review'],
            'likes': group.iloc[0]['likes'],
            'tgl_ulasan': group.iloc[0]['tgl_ulasan'],
            'thn_ulasan': group.iloc[0]['thn_ulasan']
        }        
        predicted_data = group[['record_no','predicted_topic', 'predicted_sentiment', 'count']].drop_duplicates().to_dict(orient='records')
        grouped_data.append({
                'main_data': main_data,
                'topic': predicted_data
            })    
    # Sort berdasarkan record_no descending
    grouped_data = sorted(grouped_data, key=lambda x: x['main_data']['record_no'], reverse=True)
    return render_template("place_review.html",data=grouped_data, place=place)

@app.route('/analisis_topik_ulasan')
def analisis_topik_ulasan():

    return render_template('analisis_topik_ulasan.html')

@app.route('/monev_pengaduan_tahunan', methods=["GET","POST"])
def monev_pengaduan_tahunan():
    # if request.form.get('tahun') != 'semua':
    if not request.form.get('tahun'):
        thn = 2023
    else:
        thn = int(request.form.get('tahun'))
    # thn = 2023
    df_pengaduan = pd.read_excel(dataset['file_pengaduan'])
    df_pengaduan['thn'] = df_pengaduan['tgl_pengaduan'].dt.year
    df_pengaduan = df_pengaduan.query("thn == @thn")
    df_pengaduan['bln'] = df_pengaduan['tgl_pengaduan'].dt.month

    df_bulan_pengaduan = df_pengaduan.groupby(["bln"]).size().reset_index(name='count')
    data_bulan = df_bulan_pengaduan.to_dict(orient="records")
    df_sumber_pengaduan = df_pengaduan.groupby(["bln","sumber"]).size().reset_index(name='count')
    data_sumber = df_sumber_pengaduan.to_dict(orient="records")

    total_pengaduan = df_pengaduan['record_no'].count()
    df_sumber_pengaduan_total = df_pengaduan.groupby(["sumber"]).size().reset_index(name='count')
    data_sumber_total = df_sumber_pengaduan_total.to_dict(orient="records")

    # Tabel Status Pengaduan
    df_status = df_pengaduan.groupby(["status_pengaduan"]).size().reset_index(name='count')

    # Ambil jumlah atau default ke 0 jika tidak ada record
    selesai_count = df_status.loc[df_status['status_pengaduan'] == 'selesai', 'count'].sum() if 'selesai' in df_status['status_pengaduan'].values else 0
    proses_count = df_status.loc[df_status['status_pengaduan'] == 'proses', 'count'].sum() if 'proses' in df_status['status_pengaduan'].values else 0
    tunda_count = df_status.loc[df_status['status_pengaduan'] == 'tunda', 'count'].sum() if 'tunda' in df_status['status_pengaduan'].values else 0

    # Buat dictionary
    status_selesai = {'status_pengaduan': 'selesai', 'jml': int(selesai_count)}
    status_proses = {'status_pengaduan': 'proses', 'jml': int(proses_count)}
    status_tunda = {'status_pengaduan': 'tunda', 'jml': int(tunda_count)}


    data_status = [
        status_selesai,
        status_proses,
        status_tunda
    ]

    nama_bulan = [
        "", "Januari", "Februari", "Maret", "April",
        "Mei", "Juni", "Juli", "Agustus", "September",
        "Oktober", "November", "Desember"
    ]

    return render_template(
        'monev_pengaduan_tahunan.html', 
        thn=thn, data_bulan=enumerate(data_bulan, start=1), 
        data_sumber=data_sumber, 
        data_status=data_status, 
        nama_bulan=nama_bulan,
        total_pengaduan=total_pengaduan,
        data_sumber_total=data_sumber_total
    )

@app.route('/download_tbl_01')
def download_tbl_01():
    # install library
    # pip install lxml
    # pip install html5lib
    
    login_url = host+"login"  # URL login Flask
    data_url = host+"monev_pengaduan_tahunan"  # URL data tabel
    session = requests.Session()

    # Login ke aplikasi Flask
    login_payload = {
        'username': 'admin@dara.com',  # Ganti dengan username yang valid
        'password': 'admin'   # Ganti dengan password yang valid
    }
    login_response = session.post(login_url, data=login_payload)

    if login_response.status_code != 200:
        return "Login gagal", 401

    try:
        # Ambil halaman dengan sesi login
        response = session.get(data_url)
        response.raise_for_status()  # Pastikan permintaan berhasil

        # Baca tabel dari HTML
        tables = pd.read_html(response.text, flavor="html5lib")
        if not tables:
            return "Tabel tidak ditemukan pada halaman", 404

        # Simpan tabel pertama ke file Excel
        df = tables[0]
        output_file = "output.xlsx"
        df.to_excel(output_file, index=False)

        # Kirim file Excel sebagai respons download
        return send_file(output_file, as_attachment=True)
    except Exception as e:
        print(f"Error: {e}")
        return str(e), 500



#============ list data ulasan pada fitur manajemen data
@app.route('/entry_pengaduan')
def entry_pengaduan():  
    #if 'user' not in session:
    #    return redirect(url_for('login'))
    # Ambil data dari Excel
    df = get_data_pengaduan()
    # Konversi DataFrame ke daftar untuk dikirim ke template
    data = df.to_dict(orient="records")  # Mengonversi setiap baris menjadi dictionary
    columns = df.columns.tolist()  # Mendapatkan nama kolom
    last_record = get_last_record()
    # print(last_record)
    return render_template("entry_pengaduan.html", data=data, columns=columns)

# download pengaduan menjadi file xlsx
@app.route('/download-pengaduan')
def download_pengaduan():
    
    df = pd.read_excel(dataset['file_pengaduan'])
    
    # 2. Simpan ke Excel di memory pakai openpyxl
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
    output.seek(0)

    # 3. Kirim file ke browser sebagai attachment
    return send_file(
        output,
        download_name='data_pengaduan.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/detail_pengaduan')
def detail_pengaduan():
    record_no = request.args.get('recNo')
    reviewer_id = int(request.args.get('revID'))

    df = pd.read_excel(dataset['file_pengaduan'])
    df = df.query("reviewer_id == @reviewer_id")
    data = df.to_dict(orient="records")

    print (df)

    return render_template("detail_pengaduan.html", data=data)

# fungsi hapus pengaduan
def delete_row_pengaduan(record_no):
    FILE_PATH = dataset['file_pengaduan']
    workbook = load_workbook(FILE_PATH)
    sheet = workbook.active
    row_deleted = False
    for row in sheet.iter_rows(min_row=2):  # Mulai dari baris kedua
        if row[0].value == record_no:
            sheet.delete_rows(row[0].row)
            row_deleted = True
            break
    workbook.save(FILE_PATH)
    if not row_deleted:
        raise ValueError(f"Record_no {record_no} tidak ditemukan.")

# Route untuk menghapus pengaduan
@app.route('/hapus_pengaduan', methods=["POST"])
def hapus_pengaduan():
    try:
        # Ambil data dari form
        record_no = int(request.form.get('recNo'))
        reviewer_id = int(request.form.get('revID'))

        # Hapus data dari file Excel
        delete_row_pengaduan(record_no)
        #result_message = delete_image(reviewer_id)

        # Flash pesan sukses
        # flash(f"Ulasan dengan record_no {record_no} berhasil dihapus. {result_message}", "success")
    except ValueError as e:
        # Flash pesan kesalahan jika record_no tidak ditemukan
        flash(str(e), "danger")
    except Exception as e:
        # Flash pesan kesalahan umum
        flash(f"Terjadi kesalahan: {str(e)}", "danger")
    # Redirect kembali ke halaman utama
    return redirect(url_for('entry_pengaduan'))
    
# Form tambah Data Pengaduan
@app.route('/form_entry_pengaduan')
def addPengaduan():
    record_no = get_last_record_pengaduan()+1
    return render_template("form_entry_pengaduan.html", record_no=record_no)

# Posting Data Pengaduan
@app.route('/savePengaduan',methods=['POST'])
def savePengaduan():
    last_record = get_last_record_pengaduan()
    record_no = last_record+1
    # Mengambil data dari form
    tgl_pengaduan_str = request.form.get('tgl_pengaduan')
    if not tgl_pengaduan_str:
        return {"error": "Tanggal pengaduan harus diisi."}, 400        
    # Konversi ke datetime
    try:
        tgl_pengaduan = datetime.strptime(tgl_pengaduan_str, '%Y-%m-%dT%H:%M')
        print("Tanggal berhasil dikonversi:", tgl_pengaduan)
        print("Tahun:", tgl_pengaduan.year)
        print("Bulan:", tgl_pengaduan.month)
        print("Hari:", tgl_pengaduan.day)
        print("Jam:", tgl_pengaduan.hour)
        print("Menit:", tgl_pengaduan.minute)
    except ValueError:
        print("Format tanggal tidak valid. Gunakan format YYYY-MM-DDTHH:MM.")

    thn = tgl_pengaduan.year
    reviewer_id = int(thn)*1000000+record_no
    name = request.form.get('name')
    telepon = request.form.get('telepon')
    alamat = request.form.get('alamat')
    sumber_pengaduan = request.form.get('sumber_pengaduan')
    isi_pengaduan = request.form.get('isi_pengaduan')

    tgl_proses = ''
    tgl_selesai = ''
    solusi_pengaduan = ''
    status_pengaduan = 'proses'

    # Menyiapkan data untuk disimpan
    data = [
        record_no, reviewer_id, tgl_pengaduan, tgl_proses, tgl_selesai, name, telepon, alamat, sumber_pengaduan, isi_pengaduan, isi_pengaduan, solusi_pengaduan, status_pengaduan
    ]

    # Menyimpan ke file Excel
    file_path = dataset['file_pengaduan']
    if os.path.exists(file_path):
        # Membuka file Excel dan menambahkan data
        wb = load_workbook(file_path)
        sheet = wb.active
        sheet.append(data)
        wb.save(file_path)
        wb.close()
        print('Berhasil tersimpan!')
    else:
        return "File gmaps_review.xlsx tidak ditemukan.", 404    
    print ('Berhasil tersimpan!')    
    return redirect('/entry_pengaduan')
    
@app.route('/form_update_status_pengaduan')
def form_update_status_pengaduan():
    # record_no = request.args.get('recNo')
    reviewer_id = int(request.args.get('revID'))

    df = pd.read_excel(dataset['file_pengaduan'])
    df = df.query("reviewer_id == @reviewer_id")
    data = df.to_dict(orient="records")

    print (df)

    return render_template("form_update_status_pengaduan.html", data=data)

@app.route('/update_pengaduan',methods=['POST'])
def update_pengaduan():
    reviewer_id = int(request.form.get('revID'))
    tgl_penyelesaian = request.form.get('tgl_penyelesaian')
    tgl_penyelesaian = datetime.strptime(tgl_penyelesaian, '%Y-%m-%dT%H:%M')
    solusi_pengaduan = request.form.get('solusi_pengaduan')
    status_pengaduan = 'selesai'

    # Load workbook dan pilih sheet
    file_path = dataset['file_pengaduan']
    wb = load_workbook(file_path)
    ws = wb["Sheet1"]  # Ganti dengan nama sheet Anda

    # Data yang ingin dicari dan diperbarui
    reviewer_id_target = reviewer_id
    data_update = {
        "tgl_penyelesaian": tgl_penyelesaian,  # Data untuk kolom E
        "solusi_pengaduan": solusi_pengaduan,  # Data untuk kolom L
        "status_pengaduan": status_pengaduan  # Data untuk kolom M
    }

    # Iterasi untuk mencari reviewer_id
    for row in ws.iter_rows(min_row=2, max_col=13):  # Asumsi data dimulai dari baris kedua
        if row[1].value == reviewer_id_target:  # Kolom B adalah indeks 1 (0-indexed)
            # Update kolom E (tgl_selesai)
            ws.cell(row=row[0].row, column=5, value=data_update["tgl_penyelesaian"])  # Kolom E = indeks 5
            # Update kolom L (solusi_pengaduan)
            ws.cell(row=row[0].row, column=12, value=data_update["solusi_pengaduan"])  # Kolom L = indeks 12
            # Update kolom M (status_pengaduan)
            ws.cell(row=row[0].row, column=13, value=data_update["status_pengaduan"])  # Kolom M = indeks 13
            print(f"Data pada reviewer_id {reviewer_id_target} berhasil diperbarui.")
            break
    else:
        print(f"Reviewer ID {reviewer_id_target} tidak ditemukan.")

    # Simpan workbook
    wb.save(file_path)
    wb.close()

    # return f"halaman \n{reviewer_id}\n{tgl_penyelesaian}\n{solusi_pengaduan}\n{status_pengaduan} "
    return redirect('/entry_pengaduan')
#====================================================


@app.route('/entry_ulasan')
def reviews():
    # Ambil data dari Excel
    df = get_data_ulasan()
    data = df.to_dict(orient="records")
    columns = df.columns.tolist()
    last_record = get_last_record()

    # Pagination logic
    page = request.args.get('page', 1, type=int)
    per_page = 20
    total = len(data)
    total_pages = math.ceil(total / per_page)

    start = (page - 1) * per_page
    end = start + per_page
    paginated_data = data[start:end]

    return render_template(
        "entry_ulasan.html",
        data=paginated_data,
        columns=columns,
        page=page,
        total_pages=total_pages
    )

# download ulasan menjadi file xlsx
@app.route('/download-ulasan')
def download_ulasan():
    
    df = pd.read_excel(dataset['file_ulasan'])
    
    # 2. Simpan ke Excel di memory pakai openpyxl
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
    output.seek(0)

    # 3. Kirim file ke browser sebagai attachment
    return send_file(
        output,
        download_name='data_ulasan.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# Form tambah Data Ulasan
@app.route('/form_entry_ulasan')
def addReview():
    return render_template("form_entry_ulasan.html")

# Posting Data Ulasan
@app.route('/saveUlasan',methods=['POST'])
def saveUlasan():
    last_record = get_last_record()
    record_no = last_record+1
    # Mengambil data dari form
    reviewer_id = request.form.get('reviewer_id')
    name = request.form.get('name')
    link = f'https://www.google.com/maps/contrib/{reviewer_id}?hl=id&ved=1t:31294&ictx=111'
    thumbnail = request.form.get('thumbnail') or ''  # Opsional
    # reviews = request.form.get('review')
    reviews = 0
    photos = request.form.get('photos') or ''  # Opsional
    localGuide = request.form.get('localGuide') or ''  # Opsional
    levelGuide = request.form.get('levelGuide') or ''
    rating = request.form.get('rating')
    duration = request.form.get('duration') or ''  # Opsional
    tgl_entry = datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # Tanggal saat ini
    tgl_ulasan = request.form.get('tgl_ulasan')
    snippet = request.form.get('review') or ''  # Opsional
    fixed_review = snippet
    likes = request.form.get('likes') or 0  # Opsional
    images = request.form.get('images') or ''  # Opsional
    response_from_owner = request.form.get('response_from_owner') or ''  # Opsional
    thn_ulasan = tgl_ulasan.split('-')[0] if tgl_ulasan else ''  # Tahun dari tanggal ulasan
    thn_ulasan = int(thn_ulasan)

    # Menyiapkan data untuk disimpan
    data = [
        record_no, reviewer_id, name, link, thumbnail, reviews, photos,
        localGuide, levelGuide, rating, duration, tgl_entry, tgl_ulasan, snippet, fixed_review,
        likes, images, response_from_owner, thn_ulasan
    ]

    # Mengunduh thumbnail dan menyimpan dengan nama reviewer_id.jpg
    output_folder = 'static/reviewer_thumbnail'
    if thumbnail is None or thumbnail == "":
        thumbnail = 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcTqDShlwPIXpZeHVOlDLQ8MrkAY3TDZr2OxCg&s'
    response = requests.get(thumbnail, stream=True)
    if response.status_code == 200:
        # Simpan file menggunakan reviewer_id sebagai nama file
        file_name = os.path.join(output_folder, f'{reviewer_id}.jpg')
        with open(file_name, 'wb') as f:
            f.write(response.content)
            print(f"Thumbnail untuk reviewer_id {reviewer_id} berhasil diunduh.")
    else:
        print(f"Thumbnail untuk reviewer_id {reviewer_id} gagal diunduh. Status code: {response.status_code}")
    
    # Menyimpan ke file Excel
    file_path = 'dataset/gmaps_review.xlsx'
    if os.path.exists(file_path):
        # Membuka file Excel dan menambahkan data
        wb = load_workbook(file_path)
        sheet = wb.active
        sheet.append(data)
        wb.save(file_path)
        wb.close()
        print('Berhasil tersimpan!')
    else:
        return "File gmaps_review.xlsx tidak ditemukan.", 404    
    print ('Berhasil tersimpan!')    
    return redirect('/entry_ulasan')

# Lihat ulasan satu - persatu
@app.route('/detail_ulasan',methods=['GET',"POST"])
def detail_ulasan():
    record_no = request.args.get('recNo')
    reviewer_id = request.args.get('revID')
    # reviewer_id = int(reviewer_id)
    name = request.args.get('name')
    ulasan = request.args.get('rev')
    thumbnail = request.args.get('thumbnail')

    df = pd.read_excel(dataset['file_ulasan'])
    df = df.query("reviewer_id == @reviewer_id")
    data = df.to_dict(orient="records")

    #data = {
    #    'record_no':record_no,
    #    'reviewer_id':reviewer_id,
    #    'name':name,
    #    'ulasan':ulasan,
    #    'thumbnail':thumbnail    
    #}
    
    return render_template('detail_ulasan.html', data=data)

def delete_row_by_record_no(record_no):
    FILE_PATH = dataset['file_ulasan']
    workbook = load_workbook(FILE_PATH)
    sheet = workbook.active
    row_deleted = False
    for row in sheet.iter_rows(min_row=2):  # Mulai dari baris kedua
        if row[0].value == record_no:
            sheet.delete_rows(row[0].row)
            row_deleted = True
            break
    workbook.save(FILE_PATH)
    if not row_deleted:
        raise ValueError(f"Record_no {record_no} tidak ditemukan.")

# Fungsi untuk menghapus gambar
def delete_image(reviewer_id):
    # Folder lokasi gambar
    IMAGE_FOLDER = "static/reviewer_thumbnail/"

    # Bangun path lengkap file
    file_path = os.path.join(IMAGE_FOLDER, f"{reviewer_id}.jpg")

    # Periksa apakah file ada
    if os.path.exists(file_path):
        os.remove(file_path)  # Hapus file
        return f"Gambar dengan reviewer_id {reviewer_id} berhasil dihapus."
    else:
        return f"Gambar dengan reviewer_id {reviewer_id} tidak ditemukan."

# Route untuk menghapus ulasan
@app.route('/hapus_ulasan', methods=["POST"])
def hapus_ulasan():
    try:
        # Ambil data dari form
        record_no = int(request.form.get('recNo'))
        reviewer_id = request.form.get('revID')

        # Hapus data dari file Excel dan gambar
        delete_row_by_record_no(record_no)
        result_message = delete_image(reviewer_id)

        # Flash pesan sukses
        flash(f"Ulasan dengan record_no {record_no} berhasil dihapus. {result_message}", "success")
    except ValueError as e:
        # Flash pesan kesalahan jika record_no tidak ditemukan
        flash(str(e), "danger")
    except Exception as e:
        # Flash pesan kesalahan umum
        flash(f"Terjadi kesalahan: {str(e)}", "danger")
    # Redirect kembali ke halaman utama
    return redirect(url_for('reviews'))

@app.route('/about')
def about():

    return render_template('about.html')

@app.route('/dokumentasi')
def dokumentasi():

    return render_template('dokumentasi.html')

@app.route('/faq')
def faq():
    df = pd.read_excel(dataset['file_faq'])
    data = df.to_dict(orient='records')
    return render_template('faq.html',faq=data)

@app.route('/settings')
def settings():

    return render_template('settings.html')

@app.route('/setting_pengguna')
def setting_pengguna():
    
    return render_template('settings_pengguna.html',users=users, dataset=dataset)

@app.route('/setting_unit')
def setting_unit():
    df = pd.read_excel(dataset['data_unit'])
    unit = df.to_dict(orient="records")
    return render_template('settings_unit.html',unit=unit)

@app.route('/setting_data_latih')
def setting_data_latih():
    df_latih = pd.read_excel(dataset['data_training'])
    jml_record = len(df_latih)

    df_topic = df_latih.groupby("topic").size().reset_index(name='jumlah_records')
    df_topic["persen"] = round(df_topic["jumlah_records"]/jml_record * 100,2)        
    data_latih = df_topic.to_dict(orient="records")

    df_jenis = df_latih.groupby("jenis").size().reset_index(name='jumlah_records')
    df_jenis["persen"] = round(df_jenis["jumlah_records"]/jml_record * 100,2)
    data_latih2 = df_jenis.to_dict(orient="records")

    df_sentimen = df_latih.groupby("sentimen").size().reset_index(name='jumlah_records')
    df_sentimen["persen"] = round(df_sentimen["jumlah_records"]/jml_record * 100,2)
    data_latih3 = df_sentimen.to_dict(orient="records")

    return render_template('settings_data_latih.html',users=users, dataset=dataset, data_latih=data_latih, data_latih2=data_latih2, data_latih3=data_latih3)

@app.route('/lacak_pengaduan')
def lacak_pengaduan():

    return render_template('lacak_pengaduan.html')

##########################################################################################################################
##########################################################################################################################

# API for predict
@app.route('/predict',methods=['GET',"POST"])
def predict():
    # Fungsi untuk membersihkan teks
    def clean_text(text):
        # Ubah menjadi lowercase
        text = text.lower()
        # Hapus karakter tidak diinginkan
        text = re.sub(r'[^a-z0-9\s.]', '', text)
        # Ganti tanda baca berulang dengan satu saja
        text = re.sub(r'[!?.,]{2,}', lambda m: m.group(0)[0], text)
        return text    
    # Baca file training data
    training_data = pd.read_excel(dataset['data_training'])

    # Pastikan kolom 'topic', 'sentimen', dan 'text' ada
    if 'topic' in training_data.columns and 'sentimen' in training_data.columns and 'text' in training_data.columns:
        # Preprocessing Data
        X_topic = training_data['text']
        y_topic = training_data['topic']

        X_sentiment = training_data['text']
        y_sentiment = training_data['sentimen']
        
        x_jenis = training_data['text']
        y_jenis = training_data['jenis']

        # Latih Model Naive Bayes untuk Topic
        topic_model = Pipeline([
            ('vectorizer', CountVectorizer()),  # Konversi teks menjadi fitur
            ('classifier', MultinomialNB())    # Model Naive Bayes
        ])
        topic_model.fit(X_topic, y_topic)

        # Latih Model Naive Bayes untuk Sentiment
        sentiment_model = Pipeline([
            ('vectorizer', CountVectorizer()),  # Konversi teks menjadi fitur
            ('classifier', MultinomialNB())    # Model Naive Bayes
        ])
        sentiment_model.fit(X_sentiment, y_sentiment)

        # Latih Model Naive Bayes untuk Jenis
        jenis_model = Pipeline([
            ('vectorizer', CountVectorizer()),  # Konversi teks menjadi fitur
            ('classifier', MultinomialNB())    # Model Naive Bayes
        ])
        jenis_model.fit(x_jenis, y_jenis)
        
    
    text = request.args.get('text')
    text = clean_text(text)

    proba_topic = topic_model.predict_proba([text])[0]
    label_topic = topic_model.predict([text])[0]
    kelas_topic = topic_model.named_steps['classifier'].classes_
    proba_topic_dict = dict(zip(kelas_topic, proba_topic))
    idx_max_topic = np.argmax(proba_topic)
    max_label_topic = topic_model.named_steps['classifier'].classes_[idx_max_topic]
    max_proba_topic = proba_topic[idx_max_topic]


    proba_sentiment = sentiment_model.predict_proba([text])[0]
    label_sentiment = sentiment_model.predict([text])[0]
    kelas_sentiment = sentiment_model.named_steps['classifier'].classes_
    proba_sentiment_dict = dict(zip(kelas_sentiment, proba_sentiment))
    idx_max_sentiment = np.argmax(proba_sentiment)
    max_label_sentiment = sentiment_model.named_steps['classifier'].classes_[idx_max_sentiment]
    max_proba_sentiment = proba_sentiment[idx_max_sentiment]

    proba_jenis = jenis_model.predict_proba([text])[0]
    label_jenis = jenis_model.predict([text])[0]
    kelas_jenis = jenis_model.named_steps['classifier'].classes_
    proba_jenis_dict = dict(zip(kelas_jenis, proba_jenis))
    idx_max_jenis = np.argmax(proba_jenis)
    max_label_jenis = jenis_model.named_steps['classifier'].classes_[idx_max_jenis]
    max_proba_jenis = proba_jenis[idx_max_jenis]

    results = {
    "clean_text": text,
    "jenis": {"predicted":max_label_jenis, "proba":round(max_proba_jenis * 100, 2)},
    "topic": {"predicted":max_label_topic, "proba":round(max_proba_topic * 100, 2)},
    "sentiment": {"predicted":max_label_sentiment, "proba":round(max_proba_sentiment * 100, 2)}
    } 

    return results


##################################################### RestFull API for chartJS
##############################################################################

#---------------------- Chart Ulasan -----------------------------#

@app.route('/chart_data_ulasan_pertahun')
def chart_data_ulasan_pertahun():
    df = get_data_ulasan()
    df = df.groupby('thn').size().reset_index(name='jumlah_records')
    labels = df['thn'].tolist()
    data_values = df['jumlah_records'].tolist()
    data = {
        "labels": labels,
        "datasets": [
            {
                "label": f"Ulasan pertahun",
                "backgroundColor": "rgba(75, 192, 192, 0.2)",
                "borderColor": "rgba(75, 192, 192, 1)",
                "borderWidth": 1,
                "data": data_values,
            }
        ],
    }
    return jsonify(data)

@app.route('/chart_data_rating_pertahun')
def chart_data_rating_pertahun():
    df = get_data_ulasan()    
    # Kelompokkan data berdasarkan tahun dan rating
    df = df.groupby(['thn', 'rating']).size().reset_index(name='jumlah_records')
    # Pivot data untuk mempermudah proses
    pivot_data = df.pivot(index='thn', columns='rating', values='jumlah_records').fillna(0)
    # Siapkan data untuk Chart.js
    labels = pivot_data.index.tolist()  # Tahun sebagai label
    datasets = []
    colors = [
        "rgba(255, 69, 0, 0.6)",  # Warna untuk rating 1
        "rgba(255, 165, 0, 0.6)",  # Warna untuk rating 2
        "rgba(255, 255, 0, 0.6)",  # Warna untuk rating 3
        "rgba(154, 205, 50, 0.6)",  # Warna untuk rating 4
        "rgba(75, 192, 75, 0.6)"  # Warna untuk rating 5
    ]
    for i, category in enumerate(pivot_data.columns):
        datasets.append({
            "label": f"Rating {category}",
            "data": pivot_data[category].tolist(),
            "fill": False,
            "borderColor": colors[i % len(colors)],  # Pilih warna secara berulang
            "tension": 0.1
        })
    data = {
        "labels": labels,
        "datasets": datasets
    }
    return jsonify(data)

@app.route('/chart_reviewer')
def chart_reviewer():
    df = get_data_ulasan()
    # Menghitung distribusi True dan False
    data_count = df['localGuide'].value_counts()  # Kolom 'status' berisi True/False
    labels = data_count.index.tolist()       # ['True', 'False']
    data_values = data_count.values.tolist() # [count_true, count_false]
    # Siapkan data untuk Chart.js
    data = {
        "labels": labels,
        "datasets": [
            {
                "data": data_values,
                "backgroundColor": ["rgba(75, 192, 192, 0.6)", "rgba(255, 99, 132, 0.6)"],  # Warna untuk True dan False
                "hoverBackgroundColor": ["rgba(75, 192, 192, 0.8)", "rgba(255, 99, 132, 0.8)"],  # Warna hover
            }
        ],
    }
    return jsonify(data)

@app.route('/chart_total_rating')
def chart_total_rating():
    df = get_data_ulasan()
    df = df.groupby('rating').size().reset_index(name='jumlah_records')
    # Urutkan rating dari 5 ke 1
    df = df.sort_values('rating', ascending=False)
    # Siapkan data untuk Chart.js
    labels = df['rating'].tolist()
    data_values = df['jumlah_records'].tolist()
    # Warna berdasarkan rating
    colors = {
        5: "rgba(75, 192, 75, 0.6)",      # Hijau
        4: "rgba(154, 205, 50, 0.6)",    # YellowGreen
        3: "rgba(255, 255, 0, 0.6)",     # Kuning
        2: "rgba(255, 165, 0, 0.6)",     # Oranye
        1: "rgba(255, 69, 0, 0.6)"       # Merah
    }
    background_colors = [colors[rating] for rating in labels]
    data = {
        "labels": labels,
        "datasets": [
            {
                "label": "Jumlah Ulasan per Rating",
                "backgroundColor": background_colors,
                "borderColor": background_colors,
                "borderWidth": 1,
                "data": data_values,
            }
        ],
    }
    return jsonify(data)

@app.route('/chart_data_topik_ulasan_pertahun')
def chart_data_topik_ulasan_pertahun():
    df = pd.read_csv(dataset['results_transformed_prediction_ulasan'])
    df = df.groupby(['thn','predicted_topic']).size().reset_index(name='jumlah_records')
    print (df)
    # Pivot data untuk mempermudah proses
    pivot_data = df.pivot(index='thn', columns='predicted_topic', values='jumlah_records').fillna(0)
    # Siapkan data untuk Chart.js
    labels = pivot_data.index.tolist()  # Tahun sebagai label
    datasets = []
    colors = [
        "rgba(255, 69, 0, 0.6)",  # Warna untuk rating 1
        "rgba(255, 165, 0, 0.6)",  # Warna untuk rating 2
        "rgba(255, 255, 0, 0.6)",  # Warna untuk rating 3
        "rgba(154, 205, 50, 0.6)",  # Warna untuk rating 4
        "rgba(75, 192, 75, 0.6)"  # Warna untuk rating 5
    ]
    for i, category in enumerate(pivot_data.columns):
        datasets.append({
            "label": f"{category}",
            "data": pivot_data[category].tolist(),
            "fill": False,
            "borderColor": colors[i % len(colors)],  # Pilih warna secara berulang
            "tension": 0.1
        })
    data = {
        "labels": labels,
        "datasets": datasets
    }
    return jsonify(data)

@app.route('/chart_data_sentimen_ulasan_pertahun')
def chart_data_sentimen_ulasan_pertahun():
    df = pd.read_csv(dataset['results_transformed_prediction_ulasan'])
    df = df.groupby(['thn','predicted_sentiment']).size().reset_index(name='jumlah_records')
    # print (df)
    # Pivot data untuk mempermudah proses
    pivot_data = df.pivot(index='thn', columns='predicted_sentiment', values='jumlah_records').fillna(0)
    # Siapkan data untuk Chart.js
    labels = pivot_data.index.tolist()  # Tahun sebagai label
    datasets = []
    colors = [
        "rgba(255, 69, 0, 0.6)",  # Warna untuk rating 1
        "rgba(255, 165, 0, 0.6)",  # Warna untuk rating 2
        "rgba(255, 255, 0, 0.6)",  # Warna untuk rating 3
    ]
    for i, category in enumerate(pivot_data.columns):
        datasets.append({
            "label": f"{category}",
            "data": pivot_data[category].tolist(),
            "fill": False,
            "borderColor": colors[i % len(colors)],  # Pilih warna secara berulang
            "tension": 0.1
        })
    data = {
        "labels": labels,
        "datasets": datasets
    }
    return jsonify(data)

@app.route('/chart_sentimen_ulasan_filter',methods=(['GET','POST']))
def chart_sentimen_ulasan_filter():
    thn = request.args.get('thn')
    topik = request.args.get('topik')
    sentimen = request.args.get('sentimen')

    df = pd.read_csv(dataset['results_transformed_prediction_ulasan'])
    if (thn is not None) and (thn != 'semua'):
        thn = int(thn)
        df = df.query("thn == @thn")
    if (topik is not None) and (topik != 'semua'):
        df = df.query("predicted_topic == @topik")
    if (sentimen is not None) and (sentimen != 'semua'):
        df = df.query("predicted_sentiment == @sentimen")

    df = df.groupby('predicted_sentiment').size().reset_index(name='count')

    print(thn, topik, sentimen)
    print (df)
    # Convert data to JSON
    data = {
        "labels": df['predicted_sentiment'].tolist(),
        "counts": df['count'].tolist()
    }    
    return jsonify(data)

@app.route('/chart_topik_ulasan_filter',methods=(['GET','POST']))
def chart_topik_ulasan_filter():
    thn = request.args.get('thn')
    topik = request.args.get('topik')
    sentimen = request.args.get('sentimen')

    df = pd.read_csv(dataset['results_transformed_prediction_ulasan'])
    if (thn is not None) and (thn != 'semua'):
        thn = int(thn)
        df = df.query("thn == @thn")
    if (topik is not None) and (topik != 'semua'):
        df = df.query("predicted_topic == @topik")
    if (sentimen is not None) and (sentimen != 'semua'):
        df = df.query("predicted_sentiment == @sentimen")

    df = df.groupby('predicted_topic').size().reset_index(name='count')

    print(thn, topik, sentimen)
    print (df)
    # Convert data to JSON
    data = {
        "labels": df['predicted_topic'].tolist(),
        "counts": df['count'].tolist()
    }    
    return jsonify(data)

#----------------- Chart Pengaduan ---------------------#

@app.route('/chart_data_pengaduan_pertahun')
def chart_data_pengaduan_pertahun():
    df = get_data_pengaduan()
    df = df.groupby('thn_pengaduan').size().reset_index(name='jumlah_records')
    labels = df['thn_pengaduan'].tolist()
    data_values = df['jumlah_records'].tolist()
    data = {
        "labels": labels,
        "datasets": [
            {
                "label": f"Pengaduan",
                "backgroundColor": "rgba(75, 192, 192, 0.2)",
                "borderColor": "rgba(75, 192, 192, 1)",
                "borderWidth": 1,
                "data": data_values,
            }
        ],
    }
    return jsonify(data)

@app.route('/chart_data_pengaduan_perbulan')
def chart_data_pengaduan_perbulan():
    import calendar
    import pandas as pd
    
    current_date = date.today()
    thn = current_date.year
    
    df = get_data_pengaduan()
    df = df.query("thn_pengaduan == @thn")
    
    # Buat daftar semua bulan dalam setahun
    all_months = pd.DataFrame({'bln': range(1, 13)})
    
    # Hitung jumlah records per bulan
    df = df.groupby('bln').size().reset_index(name='jumlah_records')
    
    # Gabungkan dengan daftar semua bulan
    df = pd.merge(all_months, df, on='bln', how='left')
    
    # Isi nilai NaN dengan 0 untuk bulan yang tidak memiliki data
    df['jumlah_records'] = df['jumlah_records'].fillna(0)
    
    # Konversi nama bulan menjadi string (opsional)
    df['bln'] = df['bln'].apply(lambda x: calendar.month_name[x])
    
    labels = df['bln'].tolist()
    data_values = df['jumlah_records'].tolist()
    
    data = {
        "labels": labels,
        "datasets": [
            {
                "label": f"Pengaduan",
                "backgroundColor": "rgba(75, 192, 192, 0.2)",
                "borderColor": "rgba(75, 192, 192, 1)",
                "borderWidth": 1,
                "data": data_values,
            }
        ],
    }
    return jsonify(data)


@app.route('/chart_sumber_pengaduan')
def chart_sumber_pengaduan():
    # df = pd.read_excel(dataset['file_pengaduan'])
    df = get_data_pengaduan()
    df = df.groupby('sumber').size().reset_index(name='count')
    print (df)
    # Convert data to JSON
    data = {
        "labels": df['sumber'].tolist(),
        "counts": df['count'].tolist()
    }    
    return jsonify(data)

@app.route('/chart_data_topik_pengaduan_pertahun')
def chart_data_topik_pengaduan_pertahun():
    df = pd.read_csv(dataset['results_transformed_prediction_pengaduan'])
    df = df.groupby(['thn','predicted_topic']).size().reset_index(name='jumlah_records')
    print (df)
    # Pivot data untuk mempermudah proses
    pivot_data = df.pivot(index='thn', columns='predicted_topic', values='jumlah_records').fillna(0)
    # Siapkan data untuk Chart.js
    labels = pivot_data.index.tolist()  # Tahun sebagai label
    datasets = []
    colors = [
        "rgba(255, 69, 0, 0.6)",  # Warna untuk rating 1
        "rgba(255, 165, 0, 0.6)",  # Warna untuk rating 2
        "rgba(255, 255, 0, 0.6)",  # Warna untuk rating 3
        "rgba(154, 205, 50, 0.6)",  # Warna untuk rating 4
        "rgba(75, 192, 75, 0.6)"  # Warna untuk rating 5
    ]
    for i, category in enumerate(pivot_data.columns):
        datasets.append({
            "label": f"{category}",
            "data": pivot_data[category].tolist(),
            "fill": False,
            #"borderColor": colors[i % len(colors)],  # Pilih warna secara berulang
            "tension": 0.1
        })
    data = {
        "labels": labels,
        "datasets": datasets
    }
    return jsonify(data)

@app.route('/chart_data_sumber_pengaduan_pertahun')
def chart_data_sumber_pengaduan_pertahun():
    df = get_data_pengaduan()
    df = df.groupby(['thn_pengaduan','sumber']).size().reset_index(name='jumlah_records')
    print (df)
    # Pivot data untuk mempermudah proses
    pivot_data = df.pivot(index='thn_pengaduan', columns='sumber', values='jumlah_records').fillna(0)
    # Siapkan data untuk Chart.js
    labels = pivot_data.index.tolist()  # Tahun sebagai label
    datasets = []
    colors = [
        "rgba(255, 69, 0, 0.6)",  # Warna untuk rating 1
        "rgba(255, 165, 0, 0.6)",  # Warna untuk rating 2
        "rgba(255, 255, 0, 0.6)",  # Warna untuk rating 3
        "rgba(154, 205, 50, 0.6)",  # Warna untuk rating 4
        "rgba(75, 192, 75, 0.6)"  # Warna untuk rating 5
    ]
    for i, category in enumerate(pivot_data.columns):
        datasets.append({
            "label": f"{category}",
            "data": pivot_data[category].tolist(),
            "fill": False,
            #"borderColor": colors[i % len(colors)],  # Pilih warna secara berulang
            "tension": 0.1
        })
    data = {
        "labels": labels,
        "datasets": datasets
    }
    return jsonify(data)

@app.route('/chart_topik_pengaduan_filter',methods=(['GET','POST']))
def chart_topik_pengaduan_filter():
    thn = request.args.get('thn')
    topik = request.args.get('topik')
    # status = request.args.get('status')

    df = pd.read_csv(dataset['results_transformed_prediction_pengaduan'])
    if (thn is not None) and (thn != 'semua'):
        thn = int(thn)
        df = df.query("thn == @thn")
    if (topik is not None) and (topik != 'semua'):
        df = df.query("predicted_topic == @topik")
    
    df = df.groupby('predicted_topic').size().reset_index(name='count')

    # Convert data to JSON
    data = {
        "labels": df['predicted_topic'].tolist(),
        "counts": df['count'].tolist()
    }    
    return jsonify(data)


#------ Analisis Topik Ulasan

# bar chart
@app.route('/chart_topik_ulasan')
def chart_topik_ulasan():
    df = pd.read_csv(dataset['results_transformed_prediction_ulasan'])
    df = df.groupby('predicted_topic').size().reset_index(name='jumlah_records')
    labels = df['predicted_topic'].tolist()
    data_values = df['jumlah_records'].tolist()
    data = {
        "labels": labels,
        "datasets": [
            {
                "label": f"Topik",
                "backgroundColor": "rgba(75, 192, 192, 0.2)",
                "borderColor": "rgba(75, 192, 192, 1)",
                "borderWidth": 1,
                "data": data_values,
            }
        ],
    }
    return jsonify(data)

@app.route('/chart_komposisi_ulasan')
def chart_komposisi_ulasan():
    df = pd.read_csv(dataset['results_prediction_ulasan'])
    df = df.groupby('predicted_jenis').size().reset_index(name='jumlah_records')
    labels = df['predicted_jenis'].tolist()
    data_values = df['jumlah_records'].tolist()
    data = {
        "labels": labels,
        "datasets": [
            {
                "label": f"Jenis",
                "backgroundColor": "rgba(75, 192, 192, 0.2)",
                "borderColor": "rgba(75, 192, 192, 1)",
                "borderWidth": 1,
                "data": data_values,
            }
        ],
    }
    return jsonify(data)

@app.route('/chart_analisis_sentimen_ulasan_pertahun',methods=(['GET','POST']))
def chart_analisis_sentimen_ulasan_pertahun():
    topik = request.args.get('topik')
    df = pd.read_csv(dataset['results_transformed_prediction_ulasan'])
    if (topik is not None) and (topik != 'semua'):
        df = df.query("predicted_topic == @topik")
    df = df.groupby(['thn','predicted_sentiment']).size().reset_index(name='jumlah_records')
    # print (df)
    # Pivot data untuk mempermudah proses
    pivot_data = df.pivot(index='thn', columns='predicted_sentiment', values='jumlah_records').fillna(0)
    # Siapkan data untuk Chart.js
    labels = pivot_data.index.tolist()  # Tahun sebagai label
    datasets = []
    colors = [
        "rgba(255, 69, 0, 0.6)",  # Warna untuk rating 1
        "rgba(255, 165, 0, 0.6)",  # Warna untuk rating 2
        "rgba(255, 255, 0, 0.6)",  # Warna untuk rating 3
    ]
    for i, category in enumerate(pivot_data.columns):
        datasets.append({
            "label": f"{category}",
            "data": pivot_data[category].tolist(),
            "fill": False,
            "borderColor": colors[i % len(colors)],  # Pilih warna secara berulang
            "tension": 0.1
        })
    data = {
        "labels": labels,
        "datasets": datasets
    }
    return jsonify(data)

##################################################### External Execute Batch
##############################################################################
#@app.route('/run_model')
#def run_model():
#    cmd = "python3"
#    subprocess.run([cmd,"persons.py"])
#    subprocess.run([cmd,"places.py"])
#    subprocess.run([cmd,"plot_wordcloud.py"])
#    subprocess.run([cmd,"cleaned_review_part02.py"])
#    subprocess.run([cmd ,"transform_review_prediction.py"])
#    return redirect(url_for('ulasan_lengkap'))

@app.route('/coba')
def coba():
    get_person_in_ulasan()
    get_place_in_ulasan()
    predict_topic_ulasan()
    transform_prediction_ulasan()

    get_person_in_pengaduan()
    get_place_in_pengaduan()
    predict_topic_pengaduan()
    transform_prediction_pengaduan()
    
    #return 'berhasil'
    return redirect(url_for("index"))

##################################################### ML predicted function
##############################################################################

def get_person_in_ulasan():
    # Membaca file gmaps_review.xlsx dan persons.txt
    # reviews_file = dataset['file_ulasan']
    persons_file = dataset['daftar_person']
    output_file = dataset['results_persons_ulasan']
    
    # Fungsi untuk membersihkan teks
    def clean_text(text):
        # Ubah menjadi huruf kecil
        text = text.lower()
        # Hapus tanda baca, tetapi pertahankan spasi
        text = re.sub(r"[^a-z0-9\s]", "", text)
        # Hapus spasi berlebihan
        text = re.sub(r"\s+", " ", text).strip()
        return text    
    def load_persons(file_path):
        with open(file_path, "r", encoding="utf-8") as f:
            return [line.strip() for line in f.readlines() if line.strip()]
    # Fungsi untuk mengecek jumlah kemunculan setiap nama dalam ulasan
    def count_persons_in_review(review, persons):
        person_counts = []
        for person in persons:
            # Gunakan regex untuk mencocokkan nama dengan bentuk tersambung
            pattern = rf"\b{re.escape(person)}\w*\b"
            matches = re.findall(pattern, review, re.IGNORECASE)
            count = len(matches)
            if count > 0:  # Hanya tambahkan jika ditemukan
                person_counts.append((person, count))
        return person_counts

    # ambil data ulasan
    reviews_df = get_data_ulasan()
    reviews_df['snippet'] = reviews_df['snippet'].fillna("").astype(str).apply(clean_text)

    persons = [clean_text(person) for person in load_persons(persons_file)]
    print(f"Daftar persons: {persons}")

    # Proses ulasan untuk mencari person dan jumlahnya
    results = []
    for _, row in reviews_df.iterrows():
        thn = row['thn']
        reviewer_id = row['reviewer_id']
        review_text = row['snippet']
        person_counts = count_persons_in_review(review_text, persons)
        
        for person, count in person_counts:
            results.append({
                "thn": thn,
                "reviewer_id": reviewer_id,
                "persons": person,
                "persons_count": count
            })

    # Simpan hasil ke file CSV
    results_df = pd.DataFrame(results)
    results_df.to_csv(output_file, index=False, encoding="utf-8")
    print(f"Hasil telah disimpan ke {output_file}")

def get_place_in_ulasan():
    # Membaca file gmaps_review.xlsx dan persons.txt
    # reviews_file = dataset['file_ulasan']
    places_file = dataset['daftar_place']
    output_file = dataset['results_places_ulasan']
    
    # Fungsi untuk membersihkan teks
    def clean_text(text):
        # Ubah menjadi huruf kecil
        text = text.lower()
        # Hapus tanda baca, tetapi pertahankan spasi
        text = re.sub(r"[^a-z0-9\s]", "", text)
        # Hapus spasi berlebihan
        text = re.sub(r"\s+", " ", text).strip()
        return text    
    def load_places(file_path):
        with open(file_path, "r", encoding="utf-8") as f:
            return [line.strip() for line in f.readlines() if line.strip()]
    # Fungsi untuk mengecek jumlah kemunculan setiap nama dalam ulasan
    def count_places_in_review(review, places):
        place_counts = []
        for place in places:
            # Gunakan regex untuk mencocokkan nama dengan bentuk tersambung
            pattern = rf"\b{re.escape(place)}\w*\b"
            matches = re.findall(pattern, review, re.IGNORECASE)
            count = len(matches)
            if count > 0:  # Hanya tambahkan jika ditemukan
                place_counts.append((place, count))
        return place_counts

    # ambil data ulasan
    reviews_df = get_data_ulasan()
    reviews_df['snippet'] = reviews_df['snippet'].fillna("").astype(str).apply(clean_text)

    places = [clean_text(place) for place in load_places(places_file)]
    print(f"Daftar placess: {places}")

    # Proses ulasan untuk mencari person dan jumlahnya
    results = []
    for _, row in reviews_df.iterrows():
        thn = row['thn']
        reviewer_id = row['reviewer_id']
        review_text = row['snippet']
        place_counts = count_places_in_review(review_text, places)
        
        for place, count in place_counts:
            results.append({
                "thn": thn,
                "reviewer_id": reviewer_id,
                "places": place,
                "places_count": count
            })

    # Simpan hasil ke file CSV
    results_df = pd.DataFrame(results)
    results_df.to_csv(output_file, index=False, encoding="utf-8")
    print(f"Hasil telah disimpan ke {output_file}")

def predict_topic_ulasan():
    # File paths
    input_file = dataset['file_ulasan']
    training_file = dataset['data_training']
    cleaned_file = dataset['results_cleaned_ulasan']
    prediction_file = dataset['results_prediction_ulasan']
    grouped_file = dataset['results_group_prediction_ulasan']

    # Fungsi untuk membersihkan teks
    def clean_text(text):
        # Ubah menjadi lowercase
        text = text.lower()
        # Hapus karakter tidak diinginkan
        text = re.sub(r'[^a-z0-9\s.]', '', text)
        # Ganti tanda baca berulang dengan satu saja
        text = re.sub(r'[!?.,]{2,}', lambda m: m.group(0)[0], text)
        return text

    # Baca file training data
    training_data = pd.read_excel(training_file)

    # Pastikan kolom 'topic', 'sentimen', dan 'text' ada
    if 'topic' in training_data.columns and 'sentimen' in training_data.columns and 'text' in training_data.columns:
        # Preprocessing Data
        X_topic = training_data['text']
        y_topic = training_data['topic']
        X_sentiment = training_data['text']
        y_sentiment = training_data['sentimen']
        
        x_jenis = training_data['text']
        y_jenis = training_data['jenis']

        # Latih Model Naive Bayes untuk Topic
        topic_model = Pipeline([
            ('vectorizer', CountVectorizer()),  # Konversi teks menjadi fitur
            ('classifier', MultinomialNB())    # Model Naive Bayes
        ])
        topic_model.fit(X_topic, y_topic)

        # Latih Model Naive Bayes untuk Sentiment
        sentiment_model = Pipeline([
            ('vectorizer', CountVectorizer()),  # Konversi teks menjadi fitur
            ('classifier', MultinomialNB())    # Model Naive Bayes
        ])
        sentiment_model.fit(X_sentiment, y_sentiment)

        # Latih Model Naive Bayes untuk Jenis
        jenis_model = Pipeline([
            ('vectorizer', CountVectorizer()),  # Konversi teks menjadi fitur
            ('classifier', MultinomialNB())    # Model Naive Bayes
        ])
        jenis_model.fit(x_jenis, y_jenis)

        # Baca file gmaps_review.xlsx
        df = pd.read_excel(input_file)

        # Pastikan kolom yang akan digunakan ada
        if 'reviewer_id' in df.columns and 'fixed_review' in df.columns:
            # Membuat daftar untuk menyimpan hasil
            cleaned_data = []
            prediction_data = []

            # Memecah teks snippet berdasarkan titik dan membersihkan
            for _, row in df.iterrows():
                record_no = row['record_no']
                thn_ulasan = row['thn_ulasan']
                reviewer_id = row['reviewer_id']
                snippet = row['fixed_review']
                if isinstance(snippet, str):  # Pastikan snippet adalah string
                    snippet = clean_text(snippet)  # Bersihkan teks
                    sentences = snippet.split('.')  # Pecah berdasarkan titik
                    for sentence in sentences:
                        sentence = sentence.strip()  # Hilangkan spasi di awal/akhir kalimat
                        if sentence:  # Abaikan kalimat kosong
                            cleaned_data.append({'reviewer_id': reviewer_id, 'snippet_cleaned': sentence})
                            # Prediksi topic dan sentiment untuk setiap kalimat
                            predicted_topic = topic_model.predict([sentence])[0]
                            predicted_sentiment = sentiment_model.predict([sentence])[0]
                            predicted_jenis = jenis_model.predict([sentence])[0]
                            prediction_data.append({
                                'record_no': record_no,
                                'thn': thn_ulasan,
                                'reviewer_id': reviewer_id,
                                'sentence':sentence,
                                'predicted_jenis': predicted_jenis,
                                'predicted_topic': predicted_topic,
                                'predicted_sentiment': predicted_sentiment
                            })

            # Membuat DataFrame dari hasil pembersihan
            cleaned_df = pd.DataFrame(cleaned_data)
            prediction_df = pd.DataFrame(prediction_data)

            # Membuat DataFrame untuk hasil grup
            grouped_df = (
                prediction_df.groupby(['thn','record_no','reviewer_id', 'predicted_topic', 'predicted_sentiment'])
                .size()
                .reset_index(name='count')
            )

            # Simpan hasil ke file
            cleaned_df.to_csv(cleaned_file, index=False, encoding='utf-8')
            prediction_df.to_csv(prediction_file, index=False, encoding='utf-8')
            grouped_df.to_csv(grouped_file, index=False, encoding='utf-8')

            print(f"Hasil pembersihan disimpan ke {cleaned_file}")
            print(f"Hasil prediksi disimpan ke {prediction_file}")
            print(f"Hasil grup disimpan ke {grouped_file}")
        else:
            print("Kolom 'reviewer_id' atau 'snippet' tidak ditemukan dalam file gmaps_review.xlsx")
    else:
        print("Kolom 'topic', 'sentimen', atau 'text' tidak ditemukan dalam file training_data.xls")


def transform_prediction_ulasan():
    # Baca file input
    # input_file = 'dataset/cleaned_review_prediction_group.csv'
    input_file = dataset['results_group_prediction_ulasan']

    # output_file = 'dataset/transformed_review_prediction.csv'
    output_file = dataset['results_transformed_prediction_ulasan']

    # Load dataset
    df = pd.read_csv(input_file)

    # Hapus sentimen "biasa" jika ada sentimen "positif" atau "negatif" pada topik yang sama
    def remove_biasa(group):
        if 'biasa' in group['predicted_sentiment'].values:
            if 'positif' in group['predicted_sentiment'].values or 'negatif' in group['predicted_sentiment'].values:
                return group[group['predicted_sentiment'] != 'biasa']
        return group

    df = df.groupby(['reviewer_id', 'predicted_topic'], group_keys=False).apply(remove_biasa)

    # Kelompokkan data berdasarkan reviewer_id dan predicted_topic
    grouped = df.groupby(['reviewer_id', 'predicted_topic'])

    # List untuk menyimpan indeks baris yang akan dihapus
    rows_to_drop = []
    # List untuk menyimpan baris yang diperbarui
    updated_rows = []

    for (reviewer_id, topic), group in grouped:
        # Filter baris dengan predicted_sentiment "positif" dan "negatif"
        positif = group[group['predicted_sentiment'] == 'positif']
        negatif = group[group['predicted_sentiment'] == 'negatif']
        
        if not positif.empty and not negatif.empty:
            # Ambil nilai count
            count_positif = positif['count'].values[0]
            count_negatif = negatif['count'].values[0]
            
            if count_positif > count_negatif:
                # Pilih positif, update count
                updated_row = positif.iloc[0].copy()
                updated_row['count'] = count_positif - count_negatif
                updated_rows.append(updated_row)
                rows_to_drop.extend(positif.index)
                rows_to_drop.extend(negatif.index)
            elif count_negatif > count_positif:
                # Pilih negatif, update count
                updated_row = negatif.iloc[0].copy()
                updated_row['count'] = count_negatif - count_positif
                updated_rows.append(updated_row)
                rows_to_drop.extend(positif.index)
                rows_to_drop.extend(negatif.index)
            else:
                # Jika count sama, jadikan sentimen "biasa"
                updated_row = positif.iloc[0].copy()
                updated_row['predicted_sentiment'] = 'biasa'
                updated_row['count'] = count_positif  # Nilai count tetap sama
                updated_rows.append(updated_row)
                rows_to_drop.extend(positif.index)
                rows_to_drop.extend(negatif.index)

    # Hapus baris yang terlibat konflik
    df.drop(rows_to_drop, inplace=True)

    # Tambahkan baris yang diperbarui ke dataframe
    df = pd.concat([df, pd.DataFrame(updated_rows)], ignore_index=True)

    # Simpan hasil ke file baru
    df.to_csv(output_file, index=False)
    print(f"Transformasi selesai. Data disimpan di: {output_file}")



def get_person_in_pengaduan():
    persons_file = dataset['daftar_person']
    output_file = dataset['results_persons_pengaduan']
    
    # Fungsi untuk membersihkan teks
    def clean_text(text):
        # Ubah menjadi huruf kecil
        text = text.lower()
        # Hapus tanda baca, tetapi pertahankan spasi
        text = re.sub(r"[^a-z0-9\s]", "", text)
        # Hapus spasi berlebihan
        text = re.sub(r"\s+", " ", text).strip()
        return text    
    def load_persons(file_path):
        with open(file_path, "r", encoding="utf-8") as f:
            return [line.strip() for line in f.readlines() if line.strip()]
    # Fungsi untuk mengecek jumlah kemunculan setiap nama dalam ulasan
    def count_persons_in_review(review, persons):
        person_counts = []
        for person in persons:
            # Gunakan regex untuk mencocokkan nama dengan bentuk tersambung
            pattern = rf"\b{re.escape(person)}\w*\b"
            matches = re.findall(pattern, review, re.IGNORECASE)
            count = len(matches)
            if count > 0:  # Hanya tambahkan jika ditemukan
                person_counts.append((person, count))
        return person_counts

    # ambil data ulasan
    reviews_df = get_data_pengaduan()
    reviews_df['isi_pengaduan'] = reviews_df['isi_pengaduan'].fillna("").astype(str).apply(clean_text)

    persons = [clean_text(person) for person in load_persons(persons_file)]
    print(f"Daftar persons: {persons}")

    # Proses ulasan untuk mencari person dan jumlahnya
    results = []
    for _, row in reviews_df.iterrows():
        thn = row['thn_pengaduan']
        reviewer_id = row['reviewer_id']
        review_text = row['isi_pengaduan']
        person_counts = count_persons_in_review(review_text, persons)
        
        for person, count in person_counts:
            results.append({
                "thn": thn,
                "reviewer_id": reviewer_id,
                "persons": person,
                "persons_count": count
            })

    # Simpan hasil ke file CSV
    results_df = pd.DataFrame(results)
    results_df.to_csv(output_file, index=False, encoding="utf-8")
    print(f"Hasil telah disimpan ke {output_file}")


def get_place_in_pengaduan():
    places_file = dataset['daftar_place']
    output_file = dataset['results_places_pengaduan']
    
    # Fungsi untuk membersihkan teks
    def clean_text(text):
        # Ubah menjadi huruf kecil
        text = text.lower()
        # Hapus tanda baca, tetapi pertahankan spasi
        text = re.sub(r"[^a-z0-9\s]", "", text)
        # Hapus spasi berlebihan
        text = re.sub(r"\s+", " ", text).strip()
        return text    
    def load_places(file_path):
        with open(file_path, "r", encoding="utf-8") as f:
            return [line.strip() for line in f.readlines() if line.strip()]
    # Fungsi untuk mengecek jumlah kemunculan setiap nama dalam ulasan
    def count_places_in_review(review, places):
        place_counts = []
        for place in places:
            # Gunakan regex untuk mencocokkan nama dengan bentuk tersambung
            pattern = rf"\b{re.escape(place)}\w*\b"
            matches = re.findall(pattern, review, re.IGNORECASE)
            count = len(matches)
            if count > 0:  # Hanya tambahkan jika ditemukan
                place_counts.append((place, count))
        return place_counts

    # ambil data pengaduan
    reviews_df = get_data_pengaduan()
    reviews_df['isi_pengaduan'] = reviews_df['isi_pengaduan'].fillna("").astype(str).apply(clean_text)

    places = [clean_text(place) for place in load_places(places_file)]
    print(f"Daftar persons: {persons}")

    # Proses ulasan untuk mencari person dan jumlahnya
    results = []
    for _, row in reviews_df.iterrows():
        thn = row['thn_pengaduan']
        reviewer_id = row['reviewer_id']
        review_text = row['isi_pengaduan']
        place_counts = count_places_in_review(review_text, places)
        
        for place, count in place_counts:
            results.append({
                "thn": thn,
                "reviewer_id": reviewer_id,
                "places": place,
                "places_count": count
            })

    # Simpan hasil ke file CSV
    results_df = pd.DataFrame(results)
    results_df.to_csv(output_file, index=False, encoding="utf-8")
    print(f"Hasil telah disimpan ke {output_file}")

def predict_topic_pengaduan():
    # File paths
    input_file = dataset['file_pengaduan']
    training_file = dataset['data_training']
    cleaned_file = dataset['results_cleaned_pengaduan']
    prediction_file = dataset['results_prediction_pengaduan']
    grouped_file = dataset['results_group_prediction_pengaduan']

    # Fungsi untuk membersihkan teks
    def clean_text(text):
        # Ubah menjadi lowercase
        text = text.lower()
        # Hapus karakter tidak diinginkan
        text = re.sub(r'[^a-z0-9\s.]', '', text)
        # Ganti tanda baca berulang dengan satu saja
        text = re.sub(r'[!?.,]{2,}', lambda m: m.group(0)[0], text)
        return text

    # Baca file training data
    # training_data = pd.read_excel(training_file)
    training_data = pd.read_excel(training_file)

    # Pastikan kolom 'topic', 'sentimen', dan 'text' ada
    if 'topic' in training_data.columns and 'sentimen' in training_data.columns and 'text' in training_data.columns:
        # Preprocessing Data
        X_topic = training_data['text']
        y_topic = training_data['topic']
        X_sentiment = training_data['text']
        y_sentiment = training_data['sentimen']
        
        x_jenis = training_data['text']
        y_jenis = training_data['jenis']

        # Latih Model Naive Bayes untuk Topic
        topic_model = Pipeline([
            ('vectorizer', CountVectorizer()),  # Konversi teks menjadi fitur
            ('classifier', MultinomialNB())    # Model Naive Bayes
        ])
        topic_model.fit(X_topic, y_topic)

        # Latih Model Naive Bayes untuk Sentiment
        sentiment_model = Pipeline([
            ('vectorizer', CountVectorizer()),  # Konversi teks menjadi fitur
            ('classifier', MultinomialNB())    # Model Naive Bayes
        ])
        sentiment_model.fit(X_sentiment, y_sentiment)

        # Latih Model Naive Bayes untuk Jenis
        jenis_model = Pipeline([
            ('vectorizer', CountVectorizer()),  # Konversi teks menjadi fitur
            ('classifier', MultinomialNB())    # Model Naive Bayes
        ])
        jenis_model.fit(x_jenis, y_jenis)

        # Baca file pengaduan
        # df = pd.read_excel(input_file)
        df = get_data_pengaduan()

        # Pastikan kolom yang akan digunakan ada
        if 'reviewer_id' in df.columns and 'fixed_pengaduan' in df.columns:
            # Membuat daftar untuk menyimpan hasil
            cleaned_data = []
            prediction_data = []

            # Memecah teks snippet berdasarkan titik dan membersihkan
            for _, row in df.iterrows():
                record_no = row['record_no']
                thn_pengaduan = row['thn_pengaduan']
                reviewer_id = row['reviewer_id']
                snippet = row['fixed_pengaduan']
                if isinstance(snippet, str):  # Pastikan snippet adalah string
                    snippet = clean_text(snippet)  # Bersihkan teks
                    sentences = snippet.split('.')  # Pecah berdasarkan titik
                    for sentence in sentences:
                        sentence = sentence.strip()  # Hilangkan spasi di awal/akhir kalimat
                        if sentence:  # Abaikan kalimat kosong
                            cleaned_data.append({'reviewer_id': reviewer_id, 'snippet_cleaned': sentence})
                            # Prediksi topic dan sentiment untuk setiap kalimat
                            predicted_topic = topic_model.predict([sentence])[0]
                            predicted_sentiment = sentiment_model.predict([sentence])[0]
                            predicted_jenis = jenis_model.predict([sentence])[0]
                            prediction_data.append({
                                'record_no': record_no,
                                'thn': thn_pengaduan,
                                'reviewer_id': reviewer_id,
                                'sentence':sentence,
                                'predicted_jenis': predicted_jenis,
                                'predicted_topic': predicted_topic,
                                'predicted_sentiment': predicted_sentiment
                            })

            # Membuat DataFrame dari hasil pembersihan
            cleaned_df = pd.DataFrame(cleaned_data)
            prediction_df = pd.DataFrame(prediction_data)

            # Membuat DataFrame untuk hasil grup
            grouped_df = (
                prediction_df.groupby(['thn','record_no','reviewer_id', 'predicted_topic', 'predicted_sentiment'])
                .size()
                .reset_index(name='count')
            )

            # Simpan hasil ke file
            cleaned_df.to_csv(cleaned_file, index=False, encoding='utf-8')
            prediction_df.to_csv(prediction_file, index=False, encoding='utf-8')
            grouped_df.to_csv(grouped_file, index=False, encoding='utf-8')

            print(f"Hasil pembersihan disimpan ke {cleaned_file}")
            print(f"Hasil prediksi disimpan ke {prediction_file}")
            print(f"Hasil grup disimpan ke {grouped_file}")
        else:
            print("Kolom 'reviewer_id' atau 'snippet' tidak ditemukan dalam file gmaps_review.xlsx")
    else:
        print("Kolom 'topic', 'sentimen', atau 'text' tidak ditemukan dalam file training_data.xls")

def transform_prediction_pengaduan():
    # Baca file input
    # input_file = 'dataset/cleaned_review_prediction_group.csv'
    input_file = dataset['results_group_prediction_pengaduan']

    # output_file = 'dataset/transformed_review_prediction.csv'
    output_file = dataset['results_transformed_prediction_pengaduan']

    # Load dataset
    df = pd.read_csv(input_file)

    # Hapus sentimen "biasa" jika ada sentimen "positif" atau "negatif" pada topik yang sama
    def remove_biasa(group):
        if 'biasa' in group['predicted_sentiment'].values:
            if 'positif' in group['predicted_sentiment'].values or 'negatif' in group['predicted_sentiment'].values:
                return group[group['predicted_sentiment'] != 'biasa']
        return group

    df = df.groupby(['reviewer_id', 'predicted_topic'], group_keys=False).apply(remove_biasa)

    # Kelompokkan data berdasarkan reviewer_id dan predicted_topic
    grouped = df.groupby(['reviewer_id', 'predicted_topic'])

    # List untuk menyimpan indeks baris yang akan dihapus
    rows_to_drop = []
    # List untuk menyimpan baris yang diperbarui
    updated_rows = []

    for (reviewer_id, topic), group in grouped:
        # Filter baris dengan predicted_sentiment "positif" dan "negatif"
        positif = group[group['predicted_sentiment'] == 'positif']
        negatif = group[group['predicted_sentiment'] == 'negatif']
        
        if not positif.empty and not negatif.empty:
            # Ambil nilai count
            count_positif = positif['count'].values[0]
            count_negatif = negatif['count'].values[0]
            
            if count_positif > count_negatif:
                # Pilih positif, update count
                updated_row = positif.iloc[0].copy()
                updated_row['count'] = count_positif - count_negatif
                updated_rows.append(updated_row)
                rows_to_drop.extend(positif.index)
                rows_to_drop.extend(negatif.index)
            elif count_negatif > count_positif:
                # Pilih negatif, update count
                updated_row = negatif.iloc[0].copy()
                updated_row['count'] = count_negatif - count_positif
                updated_rows.append(updated_row)
                rows_to_drop.extend(positif.index)
                rows_to_drop.extend(negatif.index)
            else:
                # Jika count sama, jadikan sentimen "biasa"
                updated_row = positif.iloc[0].copy()
                updated_row['predicted_sentiment'] = 'biasa'
                updated_row['count'] = count_positif  # Nilai count tetap sama
                updated_rows.append(updated_row)
                rows_to_drop.extend(positif.index)
                rows_to_drop.extend(negatif.index)

    # Hapus baris yang terlibat konflik
    df.drop(rows_to_drop, inplace=True)

    # Tambahkan baris yang diperbarui ke dataframe
    df = pd.concat([df, pd.DataFrame(updated_rows)], ignore_index=True)

    # Simpan hasil ke file baru
    df.to_csv(output_file, index=False)
    print(f"Transformasi selesai. Data disimpan di: {output_file}")


##############################################################################
################################## MAIN PROGRAM ##############################

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True)
